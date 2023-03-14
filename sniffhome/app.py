import os
import json
from openpyxl import load_workbook
from openpyxl import utils

from . import settings


class ConfigurationHandler:

    def __init__(self):
        self.json_obj = None
        self.workbook = None
        self.config_map = dict()        
        self.searches = dict()        

    def __del__(self):
        if self.workbook:
            self.workbook.close()            

    def load_configuration(self):
        self.__initialize_excel_config()
        self.__load_config_from_json()
        self.__map_labels_to_keys()
        self.__setup_searches()
        
    def __initialize_excel_config(self):        
        self.workbook = load_workbook(filename=os.path.join(settings.CONFIG_DIRECTORY, settings. EXCEL_CONFIG_FILE_NAME), read_only=True)
    
    def __load_config_from_json(self):
        config_file = open(os.path.abspath(os.path.join(settings.CONFIG_DIRECTORY,  settings.JSON_CONFIG_FILE_NAME)))
        self.json_obj = json.load(config_file)
        config_file.close()

    def __map_labels_to_keys(self):
        key_maps = self.json_obj["key_maps"]
        if key_maps:
            for obj_ in key_maps:
                self.config_map[obj_["key"]] = ConfigMap()
                self.config_map[obj_["key"]].key = obj_["key"]
                self.config_map[obj_["key"]].sheet = obj_["sheet"]
                self.config_map[obj_["key"]].column = obj_["column"]
                self.config_map[obj_["key"]].label = self.__get_cell_value(obj_["sheet"], obj_["label_cell"])

    def __get_cell_value(self, sheet, cell):        
        wb_sheet = self.workbook[sheet]                
        return wb_sheet[cell].value
    
    def __setup_searches(self):
        number_of_searches = 0
        reference_sheet = self.config_map["search_link"].sheet        
        ws = self.workbook[reference_sheet]        
        col_index_end= utils.column_index_from_string(self.config_map["search_status"].column)        
        for cell in ws.iter_rows(min_row=2, min_col=0, max_col=col_index_end, values_only = True):                       
            supplier_code = cell[0]
            search = Search()
            search.supplier_code = supplier_code
            search.name = cell[1]
            search.link = cell[2]
            search.use_selenium = True if cell[3] == "Yes"  else False
            search.load_wait_seconds = int(cell[4])
            search.enable = True if cell[5] == "Run"  else False
            if supplier_code not in  self.searches:
                self.searches[supplier_code] = []
            self.searches[supplier_code].append(search)
            number_of_searches += 1            
        
        attr_patterns_sheet = self.config_map["realty_container_pattern"].sheet        
        ws = self.workbook[attr_patterns_sheet] 
        col_index_end= utils.column_index_from_string(self.config_map["change_page_pattern"].column)        
        for cell in ws.iter_rows(min_row=2, min_col=0, max_col=col_index_end, values_only = True):
            supplier_code = cell[0]
            if supplier_code in self.searches:
                for search in self.searches[supplier_code]:
                    search.container_parser = cell[1]
                    search.change_page_parser = cell[2]               

        attr_patterns_sheet = self.config_map["real_state_code"].sheet        
        ws = self.workbook[attr_patterns_sheet] 
        col_index_end= utils.column_index_from_string(self.config_map["number_of_garages"].column)        
        for cell in ws.iter_rows(min_row=2, min_col=0, max_col=col_index_end, values_only = True):
            supplier_code = cell[0]
            if supplier_code in self.searches:
                for search in self.searches[supplier_code]:                    
                    search.real_state_name.parser = cell[1]
                    search.price.parser = cell[2]
                    search.condominium.parser = cell[3]
                    search.other_tax.parser = cell[4]
                    search.description.parser = cell[5]
                    search.street.parser = cell[6]
                    search.neighborhood.parser = cell[6]
                    search.area.parser = cell[7]
                    search.rooms.parser = cell[8]
                    search.garages.parser = cell[9]

        attr_patterns_sheet = self.config_map["real_state_extractor"].sheet        
        ws = self.workbook[attr_patterns_sheet] 
        col_index_end= utils.column_index_from_string(self.config_map["garage_extractor"].column)        
        for cell in ws.iter_rows(min_row=2, min_col=0, max_col=col_index_end, values_only = True):
            supplier_code = cell[0]
            if supplier_code in self.searches:
                for search in self.searches[supplier_code]:                    
                    search.real_state_name.extractor_pattern = cell[1]
                    search.price.extractor_pattern = cell[2]
                    search.condominium.extractor_pattern = cell[3]
                    search.other_tax.extractor_pattern = cell[4]
                    search.description.extractor_pattern = cell[5]
                    search.neighborhood.extractor_pattern = cell[6]
                    search.street.extractor_pattern = cell[7]
                    search.area.extractor_pattern = cell[8]
                    search.rooms.extractor_pattern = cell[9]
                    search.garages.extractor_pattern = cell[10]                
        
        
class ConfigMap:

    def __init__(self):
        self.key = ""
        self.label = ""
        self.sheet = ""
        self.column = ""
   
class RealtyFeature:

    REGEX = 1
    SPLIT = 2
    
    def __init__(self):
        self.parser = ""
        self.extractor_pattern = ""
        self.extractor_type = None
        self.split_string = ""
        self.index = 0
    
    def load_extractor_attr(self):
        pass

class Search:

    def __init__(self):
        self.supplier_code = ""
        self.name = ""
        self.link = ""
        self.use_selenium = False
        self.load_wait_seconds = 0
        self.enable = True
        self.container_parser = ""
        self.change_page_parser = ""
        self.real_state_name = RealtyFeature()
        self.price = RealtyFeature()
        self.condominium = RealtyFeature()
        self.other_tax = RealtyFeature()
        self.description = RealtyFeature()        
        self.street = RealtyFeature()
        self.neighborhood = RealtyFeature()        
        self.area = RealtyFeature()
        self.rooms = RealtyFeature()
        self.garages = RealtyFeature()

    def __repr__(self):
        return "Supplier: {supplier}\n"\
                "Name: {name}\n"\
                "Selenium: {selenium}\n"\
                "Wait Seconds: {wait}\n"\
                "Content Parser: {content}\n"\
                "Change Page Parser: {change_page}\n"\
                "Active: {active}\n"\
        .format(supplier = self.supplier_code, name=self.name, selenium = self.use_selenium, wait = self.load_wait_seconds, \
                content=self.container_parser, change_page= self.change_page_parser, active= self.enable)

    

class RealtyAddressParser:

    def __init__(self, address_parser, neighborhood_extractor, street_extractor):
        self.parser = address_parser
        self.neighborhood_extractor

class RealtyFeatures:

    def __init__(self):        
        self.description = RealtyFeature()
        self.realty_code = RealtyFeature()
        self.price = RealtyFeature()
        self.condominium = RealtyFeature()
        self.other_tax = RealtyFeature()
        self.neighborhood = RealtyFeature()
        self.street = RealtyFeature()
        self.number = RealtyFeature()
        self.area = RealtyFeature()
        self.rooms = RealtyFeature()
        self.garages = RealtyFeature()

